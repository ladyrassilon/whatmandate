from openpyxl import load_workbook
from votedata.models import *


def import_results_data(path):
	results_workbook = load_workbook(filename = path, use_iterators = True)
	results_data = results_workbook.worksheets[0]
	results_data_iterator = results_data.iter_rows()
	first_row = results_data_iterator.next()
	data_map = {}
	for column_id in range(len(first_row)):
		data_map[first_row[column_id].internal_value] = column_id

	for row in results_data_iterator:
		election_name = row[data_map['ElectionName']].internal_value
		constituency_name = row[data_map['ConstituencyName']].internal_value
		constituency_id = row[data_map['ONSConstID']].internal_value
		constituency_region_name = row[data_map['RegionName']].internal_value
		electorate_size = row[data_map['Electorate']].internal_value
		candidate_name = row[data_map['CandidateNameDisplay']].internal_value
		candidate_party = row[data_map['CandidateParty']].internal_value
		candidate_votes = row[data_map['Votes']].internal_value
		election = Election.objects.get_or_create(title=election_name)[0]
		region = Region.objects.get_or_create(name=constituency_region_name)[0]
		constituency = Constituency.objects.get_or_create(name=constituency_name,region=region)[0]
		constituency_election = ConstituencyElection.objects.get_or_create(election=election,constituency=constituency,electorate_size=electorate_size)[0]
		party = PoliticalParty.objects.get_or_create(name=candidate_party)[0]
		candidate = Candidate.objects.get_or_create(name=candidate_name,party=party,constituency_id=constituency_id)[0]
		candidate_result = CandidateResult.objects.get_or_create(candidate=candidate,constituency_election=constituency_election,votes=candidate_votes)[0]

def prime_cache():
	for constituency_election in ConstituencyElection.objects.all():
		print "turnout %s"%constituency_election.turnout_percentage()
		print "winner %s"%constituency_election.winner_total_percentage()
		print "abstention %s"%constituency_election.abstention_percentage() 